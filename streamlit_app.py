"""
Metal AM Porosity Analyzer — single-file Streamlit app.

- Deterministic OpenCV pipeline (no ML).
- Upload images or a ZIP of images.
- Adjust thresholds, review ACCEPT/REJECT, download Excel + sorted annotated images.
"""

from datetime import datetime
import json
import io
import math
import os
import zipfile
from dataclasses import dataclass, field
from typing import Dict, List, Tuple

import cv2
import numpy as np
import pandas as pd
import streamlit as st


# ------------------------------- Data models ------------------------------- #
@dataclass
class Pore:
    id: int
    centroid_x: int
    centroid_y: int
    area_pixels: float
    area_um2: float
    perimeter: float
    circularity: float
    aspect_ratio: float
    solidity: float
    equivalent_diameter_px: float
    equivalent_diameter_um: float
    pore_type: str
    contour: np.ndarray = field(repr=False)


@dataclass
class AnalysisResult:
    image_name: str
    original_image: np.ndarray
    cropped_bead: np.ndarray
    dome_mask: np.ndarray
    interior_mask: np.ndarray
    pore_mask: np.ndarray
    pore_binary_debug: np.ndarray
    annotated_image: np.ndarray
    pores: List[Pore]
    total_pores: int
    total_pore_area_px: float
    interior_area_px: float
    porosity_percentage: float
    pore_type_counts: Dict[str, int]
    scale_um_per_pixel: float
    decision: str
    quality_score: float
    confidence: float


# ------------------------- Analyzer (deterministic CV) --------------------- #
class BeadPorosityAnalyzer:
    def __init__(
        self,
        scale_um_per_pixel: float = 4.0,
        accept_porosity_max: float = 0.5,
        edge_margin_px: int = 16,
        overlay_mask: bool = True,
        baseline_offset_px: int = 0,
        fuse_thresholds: bool = True,
    ):
        self.scale = float(scale_um_per_pixel)
        self.accept_porosity_max = float(accept_porosity_max)
        self.edge_margin_px = int(edge_margin_px)
        self.overlay_mask = bool(overlay_mask)
        self.baseline_offset_px = int(baseline_offset_px)
        self.fuse_thresholds = bool(fuse_thresholds)

    @staticmethod
    def _safe_bgr(image: np.ndarray) -> np.ndarray:
        if image is None:
            return image
        if len(image.shape) == 2:
            return cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
        return image

    def _mask_overlays(self, gray: np.ndarray) -> np.ndarray:
        h, w = gray.shape
        clean = gray.copy()
        if not self.overlay_mask:
            return clean
        tr_h = int(0.22 * h)
        tr_w = int(0.30 * w)
        clean[0:tr_h, w - tr_w:w] = 0
        br_h = int(0.14 * h)
        br_w = int(0.46 * w)
        clean[h - br_h:h, w - br_w:w] = 0
        return clean

    @staticmethod
    def _fill_holes(mask_255: np.ndarray) -> np.ndarray:
        m = (mask_255 > 0).astype(np.uint8) * 255
        h, w = m.shape
        flood = m.copy()
        ff = np.zeros((h + 2, w + 2), np.uint8)
        cv2.floodFill(flood, ff, (0, 0), 255)
        flood_inv = cv2.bitwise_not(flood)
        return cv2.bitwise_or(m, flood_inv)

    @staticmethod
    def _keep_largest_component(mask_255: np.ndarray) -> np.ndarray:
        m = (mask_255 > 0).astype(np.uint8)
        n, labels, stats, _ = cv2.connectedComponentsWithStats(m, connectivity=8)
        if n <= 1:
            return mask_255
        areas = stats[1:, cv2.CC_STAT_AREA]
        idx = 1 + int(np.argmax(areas))
        return (labels == idx).astype(np.uint8) * 255

    @staticmethod
    def _illumination_normalize(gray: np.ndarray) -> np.ndarray:
        """Lightweight homomorphic-like normalization to tame hotspots."""
        gray_f = gray.astype(np.float32) + 1.0
        log = np.log(gray_f)
        blur = cv2.GaussianBlur(log, (25, 25), 0)
        norm = cv2.normalize(log - blur, None, 0, 255, cv2.NORM_MINMAX)
        return norm.astype(np.uint8)

    @staticmethod
    def _quality_metrics(gray: np.ndarray) -> tuple[float, float]:
        """Return focus (Laplacian variance) and SNR estimate."""
        focus = float(cv2.Laplacian(gray, cv2.CV_64F).var())
        mean = float(np.mean(gray))
        std = float(np.std(gray))
        snr = float(mean / (std + 1e-6))
        return focus, snr

    def extract_dome_only(self, image: np.ndarray) -> Tuple[np.ndarray, np.ndarray, np.ndarray, dict]:
        img_bgr = self._safe_bgr(image)
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
        h, w = gray.shape
        clean = self._mask_overlays(gray)

        blur = cv2.GaussianBlur(clean, (5, 5), 0)
        _, bright = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        contours, _ = cv2.findContours(bright, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            fallback = np.ones_like(gray) * 255
            return img_bgr, fallback, fallback, {"method": "fallback_no_contours"}

        main = max(contours, key=cv2.contourArea)
        material_mask = np.zeros_like(gray)
        cv2.drawContours(material_mask, [main], -1, 255, -1)

        material_mask = self._fill_holes(material_mask)
        material_mask = self._keep_largest_component(material_mask)

        tops = np.full(w, h, dtype=np.int32)
        for col in range(w):
            rows = np.where(material_mask[:, col] > 0)[0]
            if rows.size > 0:
                tops[col] = rows[0]

        valid = tops < h
        cols = np.arange(w)[valid]
        topsv = tops[valid]
        if cols.size < 250:
            fallback = material_mask.copy()
            return img_bgr, fallback, fallback, {"method": "fallback_low_profile"}

        k = min(61, max(11, int(cols.size // 20) | 1))
        if k % 2 == 0:
            k += 1
        tops_smooth = cv2.GaussianBlur(topsv.astype(np.float32).reshape(-1, 1), (k, 1), 0).ravel()

        apex_idx = int(np.argmin(tops_smooth))
        apex_col = int(cols[apex_idx])
        apex_row = float(tops_smooth[apex_idx])

        q = max(10, int(0.25 * cols.size))
        left = tops_smooth[:q]
        right = tops_smooth[-q:]
        outer = np.concatenate([left, right])

        med = float(np.median(outer))
        mad = float(np.median(np.abs(outer - med))) + 1e-6
        keep = np.abs(outer - med) < 3.5 * mad
        baseline = float(np.median(outer[keep])) if keep.any() else med
        baseline = baseline + float(self.baseline_offset_px)

        rise_threshold = 18.0
        dome_inds = np.where(tops_smooth < baseline - rise_threshold)[0]
        if dome_inds.size < 30:
            dome_left_idx = max(0, apex_idx - 60)
            dome_right_idx = min(cols.size - 1, apex_idx + 60)
        else:
            splits = np.where(np.diff(dome_inds) > 1)[0] + 1
            segs = np.split(dome_inds, splits)
            chosen = None
            for s in segs:
                if s.size and (apex_idx >= s[0] and apex_idx <= s[-1]):
                    chosen = s
                    break
            if chosen is None:
                chosen = max(segs, key=lambda s: s.size)
            dome_left_idx = int(chosen[0])
            dome_right_idx = int(chosen[-1])

        dome_left = int(cols[dome_left_idx])
        dome_right = int(cols[dome_right_idx])

        left_base = float(np.median(left))
        right_base = float(np.median(right))

        span_cols = cols[dome_left_idx:dome_right_idx + 1]
        if span_cols.size < 5:
            fallback = material_mask.copy()
            return img_bgr, fallback, fallback, {"method": "fallback_span_small"}

        t = (span_cols - span_cols[0]) / (float(span_cols[-1] - span_cols[0]) + 1e-6)
        baseline_line = (1 - t) * left_base + t * right_base
        baseline_line = baseline_line + float(self.baseline_offset_px)

        dome_mask = np.zeros_like(gray)
        for i, col in enumerate(span_cols):
            top = int(tops[col])
            base = int(baseline_line[i])
            base = max(0, min(h - 1, base))
            top = max(0, min(base, top))
            dome_mask[top:base + 1, col] = material_mask[top:base + 1, col]

        dome_mask = cv2.morphologyEx(dome_mask, cv2.MORPH_CLOSE, np.ones((11, 11), np.uint8))
        dome_mask = self._fill_holes(dome_mask)
        dome_mask = self._keep_largest_component(dome_mask)

        dist = cv2.distanceTransform((dome_mask > 0).astype(np.uint8), cv2.DIST_L2, 5)
        interior_mask = (dist > float(self.edge_margin_px)).astype(np.uint8) * 255

        ys = np.where(np.any(dome_mask > 0, axis=1))[0]
        xs = np.where(np.any(dome_mask > 0, axis=0))[0]
        if ys.size == 0 or xs.size == 0:
            fallback = material_mask.copy()
            return img_bgr, fallback, fallback, {"method": "fallback_no_dome"}

        pad = 20
        y1 = max(0, int(ys[0] - pad))
        y2 = min(h, int(ys[-1] + pad))
        x1 = max(0, int(xs[0] - pad))
        x2 = min(w, int(xs[-1] + pad))

        info = {
            "method": "baseline_profile_v5",
            "apex": (apex_col, apex_row),
            "baseline": baseline,
            "dome_left": dome_left,
            "dome_right": dome_right,
        }
        return img_bgr[y1:y2, x1:x2], dome_mask[y1:y2, x1:x2], interior_mask[y1:y2, x1:x2], info

    def detect_pores(
        self,
        image: np.ndarray,
        interior_mask: np.ndarray,
        min_area_px: int = 10,
        max_area_px: int = 8000,
        sensitivity_k: float = 2.2,
        method: str = "stats",
        watershed_split: bool = True,
        fuse_thresholds: bool | None = None,
    ) -> Tuple[List[Pore], np.ndarray, np.ndarray]:
        img_bgr = self._safe_bgr(image)
        gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)

        norm = self._illumination_normalize(gray)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        eq = clahe.apply(norm)

        se = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (13, 13))
        bh = cv2.morphologyEx(eq, cv2.MORPH_BLACKHAT, se)

        interior = (interior_mask > 0)
        pore_binary = np.zeros_like(gray, dtype=np.uint8)

        fuse = self.fuse_thresholds if fuse_thresholds is None else fuse_thresholds

        def stats_mask():
            vals = bh[interior]
            if vals.size == 0:
                return np.zeros_like(gray)
            thr = float(np.mean(vals) + sensitivity_k * np.std(vals))
            thr = max(12.0, min(thr, 90.0))
            out = np.zeros_like(gray, dtype=np.uint8)
            out[(bh > thr) & interior] = 255
            return out

        def adaptive_mask():
            tmp = bh.copy()
            tmp[~interior] = 0
            thr_img = cv2.adaptiveThreshold(
                tmp, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 35, -3
            )
            return (thr_img & interior_mask).astype(np.uint8)

        stats_bin = stats_mask()
        adaptive_bin = adaptive_mask() if method == "adaptive" or fuse else np.zeros_like(gray)

        if fuse:
            pore_binary = cv2.bitwise_or(stats_bin, adaptive_bin)
        else:
            pore_binary = adaptive_bin if method == "adaptive" else stats_bin

        pore_binary = cv2.morphologyEx(
            pore_binary, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        )
        pore_binary = cv2.morphologyEx(
            pore_binary, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        )

        final_mask = pore_binary.copy()
        if watershed_split:
            m = (pore_binary > 0).astype(np.uint8)
            dist = cv2.distanceTransform(m, cv2.DIST_L2, 5)
            if dist.max() > 0:
                dist_n = (dist / dist.max() * 255).astype(np.uint8)
                _, sure_fg = cv2.threshold(dist_n, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                sure_fg = sure_fg.astype(np.uint8)
                unknown = cv2.subtract(m * 255, sure_fg)

                n, markers = cv2.connectedComponents(sure_fg)
                markers = markers + 1
                markers[unknown > 0] = 0

                ws_img = cv2.cvtColor(eq, cv2.COLOR_GRAY2BGR)
                markers = cv2.watershed(ws_img, markers)
                final_mask = np.zeros_like(gray, dtype=np.uint8)
                final_mask[markers > 1] = 255
                final_mask[~interior] = 0

        contours, _ = cv2.findContours(final_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        pores: List[Pore] = []
        pore_mask_filled = np.zeros_like(gray, dtype=np.uint8)

        pid = 1
        for c in contours:
            area = float(cv2.contourArea(c))
            if area < float(min_area_px) or area > float(max_area_px):
                continue
            x, y, w, h = cv2.boundingRect(c)
            if w <= 0 or h <= 0:
                continue
            test = np.zeros_like(gray, dtype=np.uint8)
            cv2.drawContours(test, [c], -1, 255, -1)
            if np.any((test > 0) & (~interior)):
                continue
            per = float(cv2.arcLength(c, True))
            circ = (4.0 * math.pi * area / (per * per)) if per > 0 else 0.0
            circ = float(min(max(circ, 0.0), 1.0))
            ar = float(max(w, h) / max(1, min(w, h)))

            hull = cv2.convexHull(c)
            hull_area = float(cv2.contourArea(hull))
            solidity = float(area / hull_area) if hull_area > 0 else 0.0

            eqd_px = float(math.sqrt(4.0 * area / math.pi))
            eqd_um = float(eqd_px * self.scale)

            M = cv2.moments(c)
            if M["m00"] > 0:
                cx = int(M["m10"] / M["m00"])
                cy = int(M["m01"] / M["m00"])
            else:
                cx = int(x + w // 2)
                cy = int(y + h // 2)

            pore_type = self.classify_pore(area, circ, ar, solidity, eqd_um)
            pores.append(
                Pore(
                    id=pid,
                    centroid_x=cx,
                    centroid_y=cy,
                    area_pixels=area,
                    area_um2=area * (self.scale ** 2),
                    perimeter=per,
                    circularity=circ,
                    aspect_ratio=ar,
                    solidity=solidity,
                    equivalent_diameter_px=eqd_px,
                    equivalent_diameter_um=eqd_um,
                    pore_type=pore_type,
                    contour=c,
                )
            )
            cv2.drawContours(pore_mask_filled, [c], -1, 255, -1)
            pid += 1

        return pores, pore_mask_filled, pore_binary

    def classify_pore(self, area_px: float, circularity: float, aspect_ratio: float, solidity: float, eqd_um: float) -> str:
        if eqd_um < 20 or area_px < 35:
            return "Micro-porosity"
        if area_px > 800 and (circularity < 0.35 or solidity < 0.85) and aspect_ratio > 2.0:
            return "Lack of Fusion"
        if circularity > 0.75 and aspect_ratio < 1.6 and solidity > 0.90:
            return "Gas Porosity"
        if circularity < 0.60 or solidity < 0.90 or aspect_ratio > 2.2:
            return "Shrinkage"
        return "Gas Porosity"

    def decide_accept_reject(self, porosity_percent: float) -> str:
        return "ACCEPT" if porosity_percent <= self.accept_porosity_max else "REJECT"

    def annotate_image(
        self,
        image: np.ndarray,
        dome_mask: np.ndarray,
        interior_mask: np.ndarray,
        pores: List[Pore],
        label_mode: str = "Type+ID",
    ) -> np.ndarray:
        img = self._safe_bgr(image).copy()
        dome_cnts, _ = cv2.findContours(dome_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cv2.drawContours(img, dome_cnts, -1, (255, 200, 0), 2)
        int_cnts, _ = cv2.findContours(interior_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cv2.drawContours(img, int_cnts, -1, (0, 255, 0), 2)

        colors = {
            "Gas Porosity": (0, 0, 255),
            "Shrinkage": (255, 0, 255),
            "Lack of Fusion": (0, 140, 255),
            "Micro-porosity": (255, 200, 0),
        }
        short = {"Gas Porosity": "G", "Shrinkage": "S", "Lack of Fusion": "LOF", "Micro-porosity": "M"}

        for p in pores:
            c = colors.get(p.pore_type, (0, 255, 0))
            cv2.drawContours(img, [p.contour], -1, c, -1)
            cv2.drawContours(img, [p.contour], -1, (255, 255, 255), 1)
            if label_mode != "None":
                if label_mode == "ID":
                    txt = f"{p.id}"
                elif label_mode == "Type":
                    txt = f"{short.get(p.pore_type, '?')}"
                else:
                    txt = f"{short.get(p.pore_type, '?')}{p.id}"
                if p.area_pixels > 35 and txt:
                    x, y = p.centroid_x + 3, p.centroid_y - 3
                    (tw, th), _ = cv2.getTextSize(txt, cv2.FONT_HERSHEY_SIMPLEX, 0.45, 1)
                    cv2.rectangle(img, (x - 1, y - th - 2), (x + tw + 2, y + 2), (0, 0, 0), -1)
                    cv2.putText(img, txt, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1, cv2.LINE_AA)
        return img

    def add_corner_mark(self, image: np.ndarray, decision: str) -> np.ndarray:
        img = self._safe_bgr(image).copy()
        h, w = img.shape[:2]
        corner = min(62, w // 8, h // 8)
        color = (0, 255, 0) if decision == "ACCEPT" else (0, 0, 255)
        pts = np.array([[w - corner, 0], [w, 0], [w, corner]], np.int32)
        cv2.fillPoly(img, [pts], color)
        return img

    def analyze(
        self,
        image: np.ndarray,
        image_name: str,
        min_pore_area_px: int,
        max_pore_area_px: int,
        sensitivity_k: float,
        label_mode: str,
        pore_method: str,
        watershed_split: bool,
        fuse_thresholds: bool | None = None,
    ) -> AnalysisResult:
        cropped, dome_mask, interior_mask, _ = self.extract_dome_only(image)
        pores, pore_mask, pore_binary = self.detect_pores(
            cropped,
            interior_mask=interior_mask,
            min_area_px=min_pore_area_px,
            max_area_px=max_pore_area_px,
            sensitivity_k=sensitivity_k,
            method=pore_method,
            watershed_split=watershed_split,
            fuse_thresholds=fuse_thresholds,
        )
        interior_area = float(np.sum(interior_mask > 0))
        total_pore_area = float(sum(p.area_pixels for p in pores))
        porosity = (total_pore_area / interior_area * 100.0) if interior_area > 0 else 0.0
        decision = self.decide_accept_reject(porosity)

        counts: Dict[str, int] = {}
        for p in pores:
            counts[p.pore_type] = counts.get(p.pore_type, 0) + 1

        annotated = self.annotate_image(cropped, dome_mask, interior_mask, pores, label_mode=label_mode)
        focus, snr = self._quality_metrics(cv2.cvtColor(cropped, cv2.COLOR_BGR2GRAY))
        coverage = float(interior_area / (cropped.shape[0] * cropped.shape[1] + 1e-6))
        confidence = float(min(1.0, 0.4 + 0.00002 * focus + 0.1 * min(snr / 10.0, 1.0) + 0.3 * coverage))
        return AnalysisResult(
            image_name=image_name,
            original_image=self._safe_bgr(image),
            cropped_bead=cropped,
            dome_mask=dome_mask,
            interior_mask=interior_mask,
            pore_mask=pore_mask,
            pore_binary_debug=pore_binary,
            annotated_image=annotated,
            pores=pores,
            total_pores=len(pores),
            total_pore_area_px=total_pore_area,
            interior_area_px=interior_area,
            porosity_percentage=float(porosity),
            pore_type_counts=counts,
            scale_um_per_pixel=self.scale,
            decision=decision,
            quality_score=focus,
            confidence=confidence,
        )


# ------------------------- Reports / Downloads ----------------------------- #
def generate_excel_report(results: List[AnalysisResult], output_path: str, accept_porosity_max: float) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    accept_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    reject_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws = wb.active
    ws.title = "Summary"
    headers = [
        "#",
        "Image Name",
        "Decision",
        "Porosity (%)",
        "Interior Pores",
        "Interior Area (px)",
        "Total Pore Area (px)",
        "Quality (focus)",
        "Confidence (%)",
        "Gas",
        "Shrinkage",
        "LOF",
        "Micro",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r.image_name)
        dec = ws.cell(row=row, column=3, value=r.decision)
        dec.fill = accept_fill if r.decision == "ACCEPT" else reject_fill
        ws.cell(row=row, column=4, value=round(r.porosity_percentage, 4))
        ws.cell(row=row, column=5, value=r.total_pores)
        ws.cell(row=row, column=6, value=int(r.interior_area_px))
        ws.cell(row=row, column=7, value=round(r.total_pore_area_px, 2))
        ws.cell(row=row, column=8, value=round(r.quality_score, 2))
        ws.cell(row=row, column=9, value=round(r.confidence * 100, 2))
        ws.cell(row=row, column=10, value=r.pore_type_counts.get("Gas Porosity", 0))
        ws.cell(row=row, column=11, value=r.pore_type_counts.get("Shrinkage", 0))
        ws.cell(row=row, column=12, value=r.pore_type_counts.get("Lack of Fusion", 0))
        ws.cell(row=row, column=13, value=r.pore_type_counts.get("Micro-porosity", 0))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(60, max_len + 2)

    ws2 = wb.create_sheet("Criteria")
    ws2.cell(row=1, column=1, value="Rule").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Value").font = Font(bold=True)
    ws2.cell(row=2, column=1, value="ACCEPT if Porosity ≤")
    ws2.cell(row=2, column=2, value=f"{accept_porosity_max:.3f}%")
    ws2.cell(row=3, column=1, value="REJECT otherwise")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 18

    wb.save(output_path)
    return output_path


def create_zip_sorted_images(results: List[AnalysisResult], analyzer: BeadPorosityAnalyzer) -> bytes:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            folder = "ACCEPT" if r.decision == "ACCEPT" else "REJECT"
            img = analyzer.add_corner_mark(r.annotated_image, r.decision)
            ok, buf = cv2.imencode(".png", img)
            if not ok:
                continue
            name = os.path.splitext(r.image_name)[0] + "_analyzed.png"
            zf.writestr(f"{folder}/{name}", buf.tobytes())
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


# ------------------------------ Streamlit UI ------------------------------- #
st.set_page_config(page_title="AM Porosity Analyzer", page_icon="🔬", layout="wide")
st.markdown(
    """
    <style>
    :root {
        --bg: #0b1020;
        --card: rgba(15, 23, 42, 0.78);
        --border: rgba(148, 163, 184, 0.28);
        --text: #e2e8f0;
        --muted: #94a3b8;
    }
    body {
        background: radial-gradient(ellipse 80% 50% at 50% -20%, rgba(120,119,198,0.15), transparent),
                    radial-gradient(ellipse 60% 40% at 100% 100%, rgba(255,193,7,0.08), transparent),
                    linear-gradient(180deg, #0f0f1a 0%, #0a0a12 100%);
        color: var(--text);
    }
    .block-container { padding-top: 1rem; max-width: 1400px; }
    .stat-card {
        border: 1px solid var(--border);
        background: var(--card);
        border-radius: 16px;
        padding: 14px 16px;
    }
    .pill {
        padding: 6px 12px;
        border-radius: 999px;
        font-size: 12px;
        font-weight: 700;
        display: inline-flex;
        align-items: center;
        gap: 6px;
    }
    .pill-accept { background: rgba(16, 185, 129, 0.15); color: #34d399; border: 1px solid rgba(16, 185, 129, 0.35); }
    .pill-reject { background: rgba(239, 68, 68, 0.15); color: #f87171; border: 1px solid rgba(239, 68, 68, 0.35); }
    .section-card {
        border: 1px solid var(--border);
        background: var(--card);
        border-radius: 18px;
        padding: 18px;
    }
    .upload-card {
        border: 2px dashed rgba(148, 163, 184, 0.35);
        background: rgba(15, 23, 42, 0.6);
        border-radius: 18px;
        padding: 24px;
    }
    .legend-badge {
        padding: 6px 10px;
        border-radius: 10px;
        font-weight: 600;
        font-size: 12px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def _load_uploads(upload_mode: str):
    images: list[tuple[str, bytes]] = []
    if upload_mode == "Images":
        files = st.file_uploader(
            "Drop weld bead images", type=["jpg", "jpeg", "png", "tif", "tiff", "bmp"], accept_multiple_files=True
        )
        for f in files or []:
            images.append((f.name, f.read()))
    else:
        zip_file = st.file_uploader("Upload a ZIP folder", type=["zip"])
        if zip_file:
            with zipfile.ZipFile(zip_file, "r") as zf:
                for name in zf.namelist():
                    if name.lower().endswith((".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp")) and not name.startswith(
                        "__MACOSX"
                    ):
                        images.append((os.path.basename(name), zf.read(name)))
            if images:
                st.success(f"Found {len(images)} image(s) in ZIP")
    return images


def _render_summary(results: list[AnalysisResult]):
    accept_count = sum(1 for r in results if r.decision == "ACCEPT")
    reject_count = len(results) - accept_count
    total = len(results)
    rate = (accept_count / total * 100) if total else 0.0
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f"<div class='stat-card'><div style='font-size:13px;color:#cbd5e1;'>Total images</div><div style='font-size:26px;font-weight:700;color:white'>{total}</div></div>", unsafe_allow_html=True)
    c2.markdown(f"<div class='stat-card'><div style='font-size:13px;color:#cbd5e1;'>✅ ACCEPT</div><div style='font-size:26px;font-weight:700;color:#34d399'>{accept_count}</div></div>", unsafe_allow_html=True)
    c3.markdown(f"<div class='stat-card'><div style='font-size:13px;color:#cbd5e1;'>❌ REJECT</div><div style='font-size:26px;font-weight:700;color:#f87171'>{reject_count}</div></div>", unsafe_allow_html=True)
    c4.markdown(f"<div class='stat-card'><div style='font-size:13px;color:#cbd5e1;'>Accept rate</div><div style='font-size:26px;font-weight:700;color:#fbbf24'>{rate:.1f}%</div></div>", unsafe_allow_html=True)


def _render_results(results: list[AnalysisResult], analyzer: BeadPorosityAnalyzer, show_debug: bool):
    st.write("---")
    filters = st.columns([2, 2, 3])
    with filters[0]:
        show_only = st.selectbox("Filter", ["All", "ACCEPT only", "REJECT only"], index=0)
    with filters[1]:
        sort_by = st.selectbox("Sort by", ["Porosity (low→high)", "Porosity (high→low)"], index=0)
    with filters[2]:
        search = st.text_input("Search by filename", "")

    view = results[:]
    if show_only == "ACCEPT only":
        view = [r for r in view if r.decision == "ACCEPT"]
    elif show_only == "REJECT only":
        view = [r for r in view if r.decision == "REJECT"]
    if search.strip():
        s = search.strip().lower()
        view = [r for r in view if s in r.image_name.lower()]
    view.sort(key=lambda r: r.porosity_percentage, reverse=(sort_by == "Porosity (high→low)"))

    for r in view:
        with st.expander(f"{r.image_name} — {r.decision} — {r.porosity_percentage:.4f}% porosity", expanded=False):
            cols = st.columns([1, 2, 2, 2])
            with cols[0]:
                st.caption("Decision")
                st.metric("Status", r.decision)
                st.metric("Pores", r.total_pores)
                st.metric("Porosity %", f"{r.porosity_percentage:.4f}")
                st.metric("Quality (focus)", f"{r.quality_score:.0f}")
                st.metric("Confidence", f"{r.confidence * 100:.1f}%")
                st.caption(
                    f"G:{r.pore_type_counts.get('Gas Porosity', 0)} "
                    f"S:{r.pore_type_counts.get('Shrinkage', 0)} "
                    f"LOF:{r.pore_type_counts.get('Lack of Fusion', 0)} "
                    f"M:{r.pore_type_counts.get('Micro-porosity', 0)}"
                )
            with cols[1]:
                st.caption("Input (original)")
                st.image(cv2.cvtColor(r.original_image, cv2.COLOR_BGR2RGB), use_container_width=True)
            with cols[2]:
                st.caption("Cropped dome ROI")
                st.image(cv2.cvtColor(r.cropped_bead, cv2.COLOR_BGR2RGB), use_container_width=True)
            with cols[3]:
                st.caption("Annotated result")
                marked = analyzer.add_corner_mark(r.annotated_image, r.decision)
                st.image(cv2.cvtColor(marked, cv2.COLOR_BGR2RGB), use_container_width=True)
            if show_debug:
                st.caption("Debug masks")
                dbg = st.columns(4)
                dbg[0].image(r.dome_mask, clamp=True, use_container_width=True, caption="Dome")
                dbg[1].image(r.interior_mask, clamp=True, use_container_width=True, caption="Interior")
                dbg[2].image(r.pore_binary_debug, clamp=True, use_container_width=True, caption="Binary (pre)")
                dbg[3].image(r.pore_mask, clamp=True, use_container_width=True, caption="Final pores")


def _render_inspector(results: list[AnalysisResult], analyzer: BeadPorosityAnalyzer, show_debug: bool):
    st.write("---")
    names = [r.image_name for r in results]
    choice = st.selectbox("Pick an image", names)
    r = next(rr for rr in results if rr.image_name == choice)

    left, right = st.columns([2, 1])
    with left:
        st.caption("Annotated result")
        img = analyzer.add_corner_mark(r.annotated_image, r.decision)
        st.image(cv2.cvtColor(img, cv2.COLOR_BGR2RGB), use_container_width=True)
        if show_debug:
            st.caption("Binary pore mask (for ImageJ-like check)")
            st.image(r.pore_binary_debug, clamp=True, use_container_width=True)
    with right:
        st.caption("Metrics")
        st.write(f"**Decision:** {r.decision}")
        st.write(f"**Porosity:** {r.porosity_percentage:.4f}%")
        st.write(f"**Interior pores:** {r.total_pores}")
        st.write(f"**Interior area (px):** {int(r.interior_area_px)}")
        st.write(f"**Quality (focus):** {r.quality_score:.0f}")
        st.write(f"**Confidence:** {r.confidence * 100:.1f}%")
        st.write("**Type counts:**")
        st.write(pd.DataFrame([{"Type": k, "Count": v} for k, v in sorted(r.pore_type_counts.items(), key=lambda x: -x[1])]))
        if show_debug:
            st.caption("Masks")
            st.image(r.dome_mask, clamp=True, use_container_width=True, caption="Dome mask")
            st.image(r.interior_mask, clamp=True, use_container_width=True, caption="Interior mask")
            st.image(r.pore_mask, clamp=True, use_container_width=True, caption="Final pores")

    st.markdown("#### Pore table (interior only)")
    if r.pores:
        df = pd.DataFrame(
            [
                {
                    "ID": p.id,
                    "Type": p.pore_type,
                    "X": p.centroid_x,
                    "Y": p.centroid_y,
                    "Area(px²)": round(p.area_pixels, 2),
                    "EqD(µm)": round(p.equivalent_diameter_um, 2),
                    "Circularity": round(p.circularity, 4),
                    "AspectRatio": round(p.aspect_ratio, 3),
                    "Solidity": round(p.solidity, 3),
                }
                for p in r.pores
            ]
        )
        st.dataframe(df, use_container_width=True, height=320)
    else:
        st.info("No pores detected in interior region for this image.")


def _render_downloads(
    results: list[AnalysisResult],
    analyzer: BeadPorosityAnalyzer,
    accept_porosity_max: float,
    settings: dict,
):
    st.write("---")
    col1, col2 = st.columns(2)
    with col1:
        excel_data = None
        if results:
            from tempfile import NamedTemporaryFile

            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                generate_excel_report(results, tmp.name, accept_porosity_max=accept_porosity_max)
                excel_data = open(tmp.name, "rb").read()
                os.unlink(tmp.name)
        st.download_button(
            "📊 Download Excel report",
            data=excel_data,
            file_name=f"AM_Porosity_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            disabled=not excel_data,
        )
    with col2:
        zip_data = create_zip_sorted_images(results, analyzer) if results else None
        st.download_button(
            "📁 Download sorted images (ZIP)",
            data=zip_data,
            file_name=f"AM_Sorted_Images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            mime="application/zip",
            use_container_width=True,
            disabled=not zip_data,
        )
    audit = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "settings": settings,
        "results": [
            {
                "name": r.image_name,
                "decision": r.decision,
                "porosity": r.porosity_percentage,
                "pores": r.total_pores,
                "quality_focus": r.quality_score,
                "confidence": r.confidence,
                "counts": r.pore_type_counts,
            }
            for r in results
        ],
    }
    st.download_button(
        "🧾 Download audit JSON",
        data=json.dumps(audit, indent=2).encode("utf-8") if results else None,
        file_name=f"AM_Porosity_Audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        mime="application/json",
        use_container_width=True,
        disabled=not results,
    )
    st.info(
        "- Excel: summary + criteria (ACCEPT if porosity ≤ threshold)\n"
        "- ZIP: annotated images split into ACCEPT/REJECT\n"
        "- JSON: audit of settings + results"
    )


def main():
    header_html = f"""
<div class=\"section-card\" style=\"padding:14px 18px; border:1px solid rgba(148,163,184,0.2); margin-bottom:12px;\">
  <div style=\"display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;\">
    <div style=\"display:flex;align-items:center;gap:14px;\">
      <div style=\"display:flex;align-items:center;gap:10px;\">
        <div style=\"position:relative;\">
          <div style=\"width:54px;height:54px;border-radius:14px;background:linear-gradient(135deg,#f59e0b,#fbbf24,#f59e0b);display:flex;align-items:center;justify-content:center;box-shadow:0 10px 30px rgba(251,191,36,0.25);\"><svg viewBox='0 0 40 40' width='32' height='32' style='color:#0f172a;'><path fill='currentColor' d='M20 4 L8 18 L14 18 L10 36 L32 16 L24 16 L30 4 Z'/></svg></div>
          <div style=\"position:absolute;right:-6px;bottom:-6px;width:16px;height:16px;border-radius:50%;background:#22c55e;border:2px solid #0f172a;\"></div>
        </div>
        <div style=\"display:flex;flex-direction:column;line-height:1.2;\">
          <span style=\"font-size:11px;letter-spacing:3px;font-weight:800;color:#fbbf24;text-transform:uppercase;\">Wichita State University</span>
          <span style=\"font-size:18px;font-weight:800;color:white;\">NIAR <span style=\"color:#94a3b8;font-weight:500;\">|</span> <span style=\"background:linear-gradient(90deg,#fde68a,#fbbf24);-webkit-background-clip:text;color:transparent;\">Emerging Technologies</span></span>
          <span style=\"font-size:10px;letter-spacing:2px;color:#94a3b8;text-transform:uppercase;\">Metal AM Porosity Analyzer</span>
        </div>
      </div>
    </div>
    <div style=\"display:flex;align-items:center;gap:10px;\">
      <div style=\"padding:8px 12px;border-radius:12px;background:rgba(30,41,59,0.7);border:1px solid rgba(148,163,184,0.25);color:#cbd5e1;font-size:12px;\">
        <span style=\"color:#22c55e;font-weight:700;\">●</span> System ready · {datetime.now().strftime('%H:%M')}
      </div>
      <div style=\"padding:8px 12px;border-radius:12px;background:rgba(59,130,246,0.15);border:1px solid rgba(59,130,246,0.35);color:#bfdbfe;font-size:12px;\">Version Pro</div>
    </div>
  </div>
</div>
"""
    st.markdown(header_html, unsafe_allow_html=True)

    with st.expander('Analysis settings', expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            st.markdown('**Decision**')
            accept_porosity_max = st.number_input('ACCEPT if Porosity ≤ (%)', 0.05, 5.0, 0.50, 0.05)
            st.markdown('**Bead / Interior**')
            edge_margin = st.slider('Edge Margin (px) (removes edge pores)', 6, 50, 16)
            baseline_offset = st.slider('Baseline offset (px)', -25, 25, 0)
            overlay_mask = st.toggle('Auto-mask overlays', value=True)
            show_debug = st.toggle('Show debug masks', value=False)
        with col2:
            st.markdown('**Pore detection**')
            min_pore = st.slider('Min pore area (px)', 5, 200, 12)
            max_pore = st.slider('Max pore area (px)', 400, 40000, 8000)
            sensitivity_k = st.slider('Sensitivity k (stats mode) (higher=fewer)', 1.0, 5.0, 2.2, 0.1)
            pore_method = st.selectbox('Pore threshold method', ['stats', 'adaptive'], index=0)
            watershed_split = st.toggle('Split merged pores (watershed)', value=True)
            fuse_thresholds = st.toggle('Fuse stats + adaptive thresholds', value=True)
            st.markdown('**Labels**')
            label_mode = st.selectbox('Pore label mode', ['Type+ID', 'Type', 'ID', 'None'], index=0)

    st.markdown('### Upload & analyze')
    st.markdown("<div class='upload-card'>Drop weld bead images or a ZIP. Supports JPG, PNG, TIFF, BMP.</div>", unsafe_allow_html=True)
    upload_mode = st.radio('Upload type', ['Images', 'ZIP folder'], horizontal=True)
    uploads = _load_uploads(upload_mode)
    if not uploads:
        st.info('Upload images or a ZIP folder to start analysis.')
        return

    settings_data = {
        'accept_porosity_max': accept_porosity_max,
        'edge_margin': edge_margin,
        'baseline_offset': baseline_offset,
        'overlay_mask': overlay_mask,
        'min_pore': min_pore,
        'max_pore': max_pore,
        'sensitivity_k': sensitivity_k,
        'pore_method': pore_method,
        'watershed_split': watershed_split,
        'label_mode': label_mode,
        'fuse_thresholds': fuse_thresholds,
    }

    analyzer = BeadPorosityAnalyzer(
        scale_um_per_pixel=4.0,
        accept_porosity_max=accept_porosity_max,
        edge_margin_px=edge_margin,
        overlay_mask=overlay_mask,
        baseline_offset_px=baseline_offset,
    )

    results: list[AnalysisResult] = []
    progress = st.progress(0, text='Analyzing...')
    for idx, (file_name, file_bytes) in enumerate(uploads):
        progress.progress((idx + 1) / len(uploads), text=f'Analyzing {idx + 1}/{len(uploads)}: {file_name}')
        img = cv2.imdecode(np.asarray(bytearray(file_bytes), dtype=np.uint8), cv2.IMREAD_COLOR)
        if img is None:
            st.warning(f'Could not read {file_name}; skipping.')
            continue
        res = analyzer.analyze(
            img,
            image_name=file_name,
            min_pore_area_px=min_pore,
            max_pore_area_px=max_pore,
            sensitivity_k=sensitivity_k,
            label_mode=label_mode,
            pore_method=pore_method,
            watershed_split=watershed_split,
            fuse_thresholds=fuse_thresholds,
        )
        results.append(res)
    progress.empty()

    if not results:
        st.error('No valid images were processed.')
        return

    _render_summary(results)
    st.markdown('---')
    tab1, tab2, tab3 = st.tabs(['Visual report', 'Inspector', 'Downloads'])
    with tab1:
        _render_results(results, analyzer, show_debug)
    with tab2:
        _render_inspector(results, analyzer, show_debug)
    with tab3:
        _render_downloads(results, analyzer, accept_porosity_max, settings_data)

    st.markdown('---')
    st.markdown(
        """
<div class=\"section-card\">
  <strong>Pore type legend</strong><br>
  <span class=\"legend-badge\" style=\"color:#f87171;background:rgba(248,113,113,0.12);\">G</span> Gas ·
  <span class=\"legend-badge\" style=\"color:#e879f9;background:rgba(232,121,249,0.12);\">S</span> Shrinkage ·
  <span class=\"legend-badge\" style=\"color:#fb923c;background:rgba(251,146,60,0.12);\">LOF</span> Lack of Fusion ·
  <span class=\"legend-badge\" style=\"color:#fbbf24;background:rgba(251,191,36,0.12);\">M</span> Micro-porosity
</div>
""",
        unsafe_allow_html=True,
    )
